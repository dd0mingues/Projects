from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import Workbook
from openpyxl import load_workbook
import openpyxl
import copy
import pandas as pd
from bs4 import BeautifulSoup
import requests

def getCollection():
    ##connect to iclintz website
    connection_url = 'https://elowki-events.iclintz.com/help/connect.php'

    driver = webdriver.Firefox()
    driver.get(connection_url)

    wait = WebDriverWait(driver, 100)
    connect_button = wait.until(EC.presence_of_element_located((By.CLASS_NAME, "color1")))
    connect_button.click()

    cookies_button = wait.until(EC.presence_of_element_located((By.CLASS_NAME, "js-urmodal-validate")))
    cookies_button.click()

    auth_button = wait.until(EC.presence_of_element_located((By.CLASS_NAME, "btn-success")))
    auth_button.click()

    event_url = 'https://elowki-events.iclintz.com/event/rke/'
    driver.get(event_url)

    progress_button = driver.find_elements(By.CLASS_NAME, 'color1')
    progress_button[0].click()

    collecntion_tab = driver.find_element(By.ID, "home")

    with open('collection.txt', 'w') as file:
        file.write(collecntion_tab.text)

    driver.close()
    return

def parseCollection():
    
    collection_content = open('collection.txt', 'r')
    char_paste_content = open('charPaste.txt', 'r')
    starter_char_paste_content = open('starterPackPaste.txt', 'r')


    if collection_content == '':
        print('collection.txt is empty')
        return

    collection_items = collection_content.read().split('\n')
    all_chracters = char_paste_content.read().split(', ')
    all_starter_chracters = starter_char_paste_content.read().split(', ')


    filename = 'MyChars.xlsx'
    wb = load_workbook(filename)
    my_collection_worksheet = overwriteExcel(wb, 'My Collenction')
    my_collection_worksheet.append(('Card Name', 'Nr Copies'))

    all_my_characters = []
    my_collection_worksheet.cell(1,1,'Card')
    my_collection_worksheet.cell(1,2,'Nr Copies')
    cell_index = 2
    for index, item in enumerate(collection_items):
        if item.isnumeric():
            continue 
        if not item.isnumeric() and collection_items[index+1].isnumeric():
            card = (item, int(collection_items[index +1]))
            all_my_characters.append(item)
            my_collection_worksheet.cell(cell_index, 1,item)
            my_collection_worksheet.cell(cell_index, 2, int(collection_items[index +1]))
            cell_index += 1

    all_my_characters_list = set(all_my_characters)
    all_characters_list = set(all_chracters)
    all_starter_chracters_list = set(all_starter_chracters)

    missing_characters = all_characters_list-all_my_characters_list
    missing_starter_characters = all_starter_chracters_list-all_my_characters_list

    missing_characters_worksheet = overwriteExcel(wb, 'Missing Characters')
    missing_starter_characters_worksheet = overwriteExcel(wb, 'Missing Starter Characters')


    for index, item in enumerate(missing_characters):
        missing_characters_worksheet.cell(index+2, 1, item)
    missing_characters_worksheet['A1'] = 'Missing Cards'
    missing_characters_worksheet['A1'].font = openpyxl.styles.Font(bold=True)

    for index, item in enumerate(missing_starter_characters):
        missing_starter_characters_worksheet.cell(index+2, 1, item)
    missing_starter_characters_worksheet['A1'] = 'Missing Cards'
    missing_starter_characters_worksheet['A1'].font = openpyxl.styles.Font(bold=True)
    

    wb.save('MyChars.xlsx')

def getCharsId():
    filename = 'MyChars.xlsx'
    url = 'https://iclintz.com/characters/card.php?ID='

    char_paste_content = open('charPaste.txt', 'r')
    usable_characters = char_paste_content.read().split(', ')

    book = load_workbook(filename)
    characters_id_worksheet = overwriteExcel(book, 'Characters ID')
    characters_id_worksheet.append(('Clan', 'Stars', 'Name', 'Pow', 'Dmg', 'Ability', 'Bonus'))

    number_of_chars = 2400
    number_array = list(range(123, number_of_chars)) #123 seems to be the first card id

    for n in number_array:
        print
        char_url = url + str(n)
        try:
            response = requests.get(char_url)
        except Exception as e:
            continue
        soup = BeautifulSoup(response.text, 'html.parser')

        h1_elements = soup.find_all('h1')
        if h1_elements:
            cardFrame_elements = soup.find_all('div', class_='cardFrame')
            for cardFrame in cardFrame_elements:
                card_name = cardFrame.find('span', class_="cardName").text
                if card_name in usable_characters:
                    card_ability = cardFrame.find('div', class_='cardPower').text.strip()
                    if 'Ability at' in card_ability:
                        continue
                    if 'No Ability' in card_ability and cardFrame.find('div', class_='cardStarOff'):
                        continue
                    
                    card_rarity = cardFrame.attrs['class'][3].split('_')[1]
                    card_stars = len(cardFrame.find_all('div', class_='cardStarOn'))
                    card_power = cardFrame.find('div', class_='cardPH').text
                    card_dmg = cardFrame.find('div', class_='cardPDD').text
                    card_bonus = cardFrame.find('div', class_='cardBonus').text.strip()
                    card_clan = cardFrame.find('img', class_='cardClanPict')['src'].split('/clan/')[1].split('_')[0]

                    ban = ""
                    page_title = soup.find('h1')
                    title_images = page_title.find_all('img')
                    if title_images:
                        for image in title_images:
                            print(image['src'].split('icon-ban-')[1].split('.')[0])
                            ban += image['src'].split('icon-ban-')[1].split('.')[0] + ' '
            
                    card_data = (card_clan,card_rarity ,card_stars, card_name, int(card_power), int(card_dmg), card_ability, card_bonus, ban, n)
                    characters_id_worksheet.append(card_data)
                    print(card_data)

            if card_name in usable_characters:
                usable_characters.remove(card_name)
    
        if (n % 25) == 0:
            print(n)
                
    print(usable_characters)
    book.save('MyChars.xlsx')

def individualCharId(id):
    url = 'https://iclintz.com/characters/card.php?ID=' + str(id)

    filename = 'MyChars.xlsx'
    book = load_workbook(filename)
    characters_id_worksheet = book['Characters ID']

    char_paste_content = open('charPaste.txt', 'r')
    usable_characters = char_paste_content.read().split(', ')

    try:
         response = requests.get(url)
    except Exception as e:
        print('failed to get')
    soup = BeautifulSoup(response.text, 'html.parser')

    h1_elements = soup.find_all('h1')
    if h1_elements:
        cardFrame_elements = soup.find_all('div', class_='cardFrame')
        for cardFrame in cardFrame_elements:
            card_name = cardFrame.find('span', class_="cardName").text
            if card_name in usable_characters:
                card_ability = cardFrame.find('div', class_='cardPower').text.strip()
                if 'Ability at' in card_ability:
                    continue
                if 'No Ability' in card_ability and cardFrame.find('div', class_='cardStarOff'):
                    continue

                card_stars = len(cardFrame.find_all('div', class_='cardStarOn'))
                card_power = cardFrame.find('div', class_='cardPH').text
                card_dmg = cardFrame.find('div', class_='cardPDD').text
                card_bonus = cardFrame.find('div', class_='cardBonus').text.strip()
                card_clan = cardFrame.find('img', class_='cardClanPict')['src'].split('/clan/')[1].split('_')[0]

                ban = ""
                page_title = soup.find('h1')
                title_images = page_title.find_all('img')
                if title_images:
                    for image in title_images:
                        ban += image['src'].split('icon-ban-')[1].split('.')[0]

            
                card_data = (card_clan, card_stars, card_name, int(card_power), int(card_dmg), card_ability, card_bonus, ban, id)
                characters_id_worksheet.append(card_data)
                print(card_data)
    book.save('MyChars.xlsx')
    



def overwriteExcel(book, sheet_name):
    if sheet_name not in book.sheetnames:
        current_worksheet = book.create_sheet(sheet_name)
    else:
        current_worksheet = book[sheet_name]
        for row in current_worksheet.iter_rows():
            for cell in row:
                cell.value = None
    
    return current_worksheet


